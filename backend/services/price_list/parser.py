from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path


PREVIEW_ROWS_COUNT = 10
REQUIRED_MAPPING_FIELDS = ("name", "price")
OPTIONAL_MAPPING_FIELDS = ("article", "unit")


class PriceListParsingError(Exception):
    pass


def _load_openpyxl():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise PriceListParsingError(
            "Excel preview requires openpyxl. Install project dependencies first."
        ) from exc

    return load_workbook


def _load_xlrd():
    try:
        import xlrd
    except ModuleNotFoundError as exc:
        raise PriceListParsingError(
            "XLS preview requires xlrd. Install project dependencies first."
        ) from exc

    return xlrd


def _column_letter(index: int) -> str:
    index += 1
    result = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def _column_index(letter: str) -> int:
    if not letter:
        raise PriceListParsingError("Column letter cannot be empty.")

    normalized = letter.strip().upper()
    if not normalized.isalpha():
        raise PriceListParsingError(f"Invalid column letter: {letter}")

    value = 0
    for char in normalized:
        value = value * 26 + (ord(char) - 64)
    return value - 1


def _normalize_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _read_workbook(file_path: str | Path):
    suffix = Path(file_path).suffix.lower()
    if suffix == ".xlsx":
        load_workbook = _load_openpyxl()
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        rows_by_sheet = {
            sheet_name: [tuple(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
            for sheet_name in sheet_names
        }
        return sheet_names, rows_by_sheet

    if suffix == ".xls":
        xlrd = _load_xlrd()
        workbook = xlrd.open_workbook(file_path)
        sheet_names = workbook.sheet_names()
        rows_by_sheet = {}
        for sheet_name in sheet_names:
            sheet = workbook.sheet_by_name(sheet_name)
            rows_by_sheet[sheet_name] = [tuple(sheet.row_values(row_idx)) for row_idx in range(sheet.nrows)]
        return sheet_names, rows_by_sheet

    raise PriceListParsingError("Only .xlsx and .xls files are supported in this version.")


def build_preview(file_path: str | Path, sheet_name: str | None = None, header_row: int = 1) -> dict:
    available_sheets, rows_by_sheet = _read_workbook(file_path)

    if not available_sheets:
        raise PriceListParsingError("Excel file does not contain sheets.")

    selected_sheet = sheet_name or available_sheets[0]
    if selected_sheet not in available_sheets:
        raise PriceListParsingError(f"Sheet '{selected_sheet}' was not found.")

    rows = rows_by_sheet[selected_sheet]

    if not rows:
        raise PriceListParsingError("Excel file is empty.")

    if header_row < 1 or header_row > len(rows):
        raise PriceListParsingError("Header row is out of bounds.")

    header_index = header_row - 1
    header_values = rows[header_index]
    columns = [
        {
            "index": idx,
            "letter": _column_letter(idx),
            "header": _normalize_cell(value) or _column_letter(idx),
        }
        for idx, value in enumerate(header_values)
    ]

    preview_rows = []
    for offset, row in enumerate(rows[header_index + 1 : header_index + 1 + PREVIEW_ROWS_COUNT], start=1):
        preview_rows.append(
            {
                "row_number": header_row + offset,
                "values": {
                    _column_letter(idx): _normalize_cell(value)
                    for idx, value in enumerate(row)
                },
            }
        )

    return {
        "sheet_names": available_sheets,
        "selected_sheet": selected_sheet,
        "header_row": header_row,
        "columns": columns,
        "rows": preview_rows,
    }


def validate_mapping(mapping: dict, available_columns: list[dict]) -> dict:
    if not isinstance(mapping, dict):
        raise PriceListParsingError("Mapping must be an object.")

    sheet = mapping.get("sheet")
    header_row = mapping.get("header_row", 1)
    fields = mapping.get("fields")

    if not sheet:
        raise PriceListParsingError("Mapping must include 'sheet'.")

    if not isinstance(header_row, int) or header_row < 1:
        raise PriceListParsingError("'header_row' must be a positive integer.")

    if not isinstance(fields, dict):
        raise PriceListParsingError("Mapping must include 'fields'.")

    available_letters = {column["letter"] for column in available_columns}

    for field_name in REQUIRED_MAPPING_FIELDS:
        if not fields.get(field_name):
            raise PriceListParsingError(f"Field '{field_name}' is required in mapping.")

    normalized_fields = {}
    for field_name in (*REQUIRED_MAPPING_FIELDS, *OPTIONAL_MAPPING_FIELDS):
        column_letter = fields.get(field_name)
        if not column_letter:
            continue

        normalized_letter = column_letter.strip().upper()
        if normalized_letter not in available_letters:
            raise PriceListParsingError(
                f"Column '{normalized_letter}' is not present in preview columns."
            )
        normalized_fields[field_name] = normalized_letter

    return {
        "sheet": sheet,
        "header_row": header_row,
        "fields": normalized_fields,
    }


def build_items_payload(file_path: str | Path, mapping: dict) -> list[dict]:
    _, rows_by_sheet = _read_workbook(file_path)
    rows = rows_by_sheet[mapping["sheet"]]
    header_row = mapping["header_row"]
    field_columns = {
        field_name: _column_index(column_letter)
        for field_name, column_letter in mapping["fields"].items()
    }

    items = []
    for row in rows[header_row:]:
        name_value = _normalize_cell(
            row[field_columns["name"]] if field_columns["name"] < len(row) else None
        )
        price_value = _normalize_cell(
            row[field_columns["price"]] if field_columns["price"] < len(row) else None
        )

        if not name_value and not price_value:
            continue

        try:
            price = Decimal(price_value.replace(",", ".")) if price_value else Decimal("0")
        except InvalidOperation as exc:
            raise PriceListParsingError(f"Invalid price value: {price_value}") from exc

        item_payload = {
            "article": _normalize_cell(
                row[field_columns["article"]] if "article" in field_columns and field_columns["article"] < len(row) else None
            ),
            "name": name_value,
            "unit": _normalize_cell(
                row[field_columns["unit"]] if "unit" in field_columns and field_columns["unit"] < len(row) else None
            ),
            "price": price,
            "raw_data": {
                _column_letter(idx): _normalize_cell(value)
                for idx, value in enumerate(row)
            },
        }
        items.append(item_payload)

    return items
